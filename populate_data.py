import openpyxl

# Import the spreadsheet to be updated
first_spreadsheet = openpyxl.load_workbook("example_file1.xlsx")
first_worksheet = first_spreadsheet.active

# Import the spreadsheet containing the data to be transferred 
second_spreadsheet = openpyxl.load_workbook("example_export.xlsx")
second_worksheet = second_spreadsheet.active

# Define column indexes for "Home #" and "Street Name & Notes"
home_column_indexes = [3, 11]  # Columns C and K
street_column_indexes = [6, 14]  # Columns F and N

# Keep track of populated street numbers
populated_street_numbers = set()

# Iterate over the rows in the second spreadsheet, skipping the header
for row_idx, row in enumerate(second_worksheet.iter_rows(min_row=2, values_only=True), start=2):
    # Get the street number from the "StreetNo" column (Column B)
    street_number = str(row[1])

    # Check if the street number has been populated before
    if street_number in populated_street_numbers:
        continue  # Skip this row

    # Populate the "Home #" columns
    for col_idx in home_column_indexes:
        first_worksheet.cell(row=row_idx, column=col_idx).value = street_number

    # Combine the street name, street type, and street suffix (Columns E, F, and G)
    street_name_parts = [str(row[i]) for i in range(4, 7) if row[i] is not None]
    street_name = " ".join(street_name_parts)

    # Populate the "Street Name & Notes" columns
    for col_idx in street_column_indexes:
        first_worksheet.cell(row=row_idx, column=col_idx).value = street_name

    # Mark the street number as populated
    populated_street_numbers.add(street_number)

# Save the new spreadsheet with the populated data
first_spreadsheet.save("example_file1.xlsx")
